"""Extract tables from Markdown, associate each table with its component
heading, filter the rows that carry a Part Number, and export everything to an
Excel workbook (one worksheet per component table).

This module is deliberately dependency-light: it only needs ``openpyxl`` (for
the Excel export).  Everything else is the standard library, so the parsing
helpers can be reused later from the Reflex web app.

Why "component"?
----------------
PDF -> Markdown conversion of parts manuals usually yields one or more tables
per machine component/sub-assembly.  The Markdown heading(s) that precede a
table describe which component the parts in the table belong to.  Each table is
therefore tagged with that heading (its "component"), and every exported row
can be traced back to the component that uses it.

Heuristics used on real (messy) converter output
-------------------------------------------------
* Tables are groups of consecutive lines starting with ``|``.  Separator rows
  (``|---|``) are ignored.
* A header may span several rows (a stray first line + the real column header),
  so consecutive "header-like" rows are merged column by column.
* Some converters merge the *column header* with the *first data row* into a
  single line (e.g. ``| Item # 1 2 | Part Number 410001183677 410001183678 | … |``).
  When a header cell contains a known column label followed by data, the line
  is split back into "labels" (the header) + "data" (the first row).
* Multi-line / wrapped cell content: when a part's text flows over several
  visual rows that carry no Part Number, those rows are merged back into the
  part that owns them (so quantities and descriptions are not lost).
* If a Part Number cell still contains several space-separated part numbers
  (several parts collapsed into one line), it is expanded into one row per part
  number on a best-effort basis and a warning is recorded.
"""

from __future__ import annotations

import html
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Small text helpers
# --------------------------------------------------------------------------

_HTML_TAG_RE = re.compile(r"<[^>]+>")
_SPAN_RE = re.compile(r"<span[^>]*>.*?</span>", re.DOTALL)
_WS_RE = re.compile(r"[ \t\r\n]+")
_MD_EMPH_RE = re.compile(r"(?<!\*)\*{1,2}(?!\*)")


def clean_text(raw: object) -> str:
    """Strip HTML tags / markdown emphasis, collapse whitespace, trim."""
    if raw is None:
        return ""
    s = str(raw)
    s = _SPAN_RE.sub(" ", s)
    s = _HTML_TAG_RE.sub(" ", s)
    s = _MD_EMPH_RE.sub("", s)
    s = html.unescape(s)
    s = _WS_RE.sub(" ", s)
    return s.strip()


# --------------------------------------------------------------------------
# Part-number detection
# --------------------------------------------------------------------------

_PN_STRIP = re.compile(r"/[A-Za-z0-9]+$")
_ALLOWED_PN_CHARS = frozenset(
    "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-"
)

_PN_PATTERNS = [
    re.compile(r"^[0-9]{6,15}$"),                    # 410001183677, 2787003
    re.compile(r"^[0-9]{4,6}[A-Za-z]$"),             # 058773M, 003088V
    re.compile(r"^[A-Za-z]{1,3}[0-9]{4,8}[A-Za-z]{0,3}$"),  # A01092Y, P42330L
    re.compile(r"^[A-Za-z]{0,3}[0-9]{2,6}-[A-Za-z0-9]{1,12}$"),  # A02-0003K
]


def is_part_number_token(token: str) -> bool:
    """Return True when ``token`` (a single word) looks like a part number."""
    # note: '.' is intentionally NOT stripped, so TOC entries padded with dots
    # like "X22124F/00........" are rejected by the allowed-chars check below.
    t = token.strip(" \t,;:()\"'`")
    if not t or len(t) < 4:
        return False
    # drop a trailing "/00"-style revision suffix ("A09082N/00" -> "A09082N")
    t = _PN_STRIP.sub("", t)
    if not t:
        return False
    # part numbers are made of letters/digits/dashes only; anything else (e.g.
    # "X22124F/00......" TOC entries, "3.5m" measures) is not a part number.
    if any(ch not in _ALLOWED_PN_CHARS for ch in t):
        return False
    for pat in _PN_PATTERNS:
        if pat.match(t):
            return True
    return False


def pn_tokens_in(text: object) -> list[str]:
    """Return every space separated token of ``text`` that looks like a part number."""
    return [tok for tok in str(text or "").split() if is_part_number_token(tok)]


def has_part_number(text: object) -> bool:
    return bool(pn_tokens_in(text))


# --------------------------------------------------------------------------
# Header label vocabulary
# --------------------------------------------------------------------------

# (display label, canonical group).  Order matters: longest first so that
# "Item #" is matched before "Item" and "Part Number" before "Part".
_HEADER_VOCAB = [
    ("Item #", "item"),
    ("Item no.", "item"),
    ("Item no", "item"),
    ("Item#", "item"),
    ("Item", "item"),
    ("No.", "item"),
    ("No", "item"),
    ("Part Number", "partnumber"),
    ("Part no.", "partnumber"),
    ("Part no", "partnumber"),
    ("Part#", "partnumber"),
    ("Part #", "partnumber"),
    ("Référence", "partnumber"),
    ("Reference", "partnumber"),
    ("Code", "partnumber"),
    ("Réf.", "partnumber"),
    ("Réf", "partnumber"),
    ("Ref.", "partnumber"),
    ("Ref", "partnumber"),
    ("PN", "partnumber"),
    ("Désignation", "designation"),
    ("Designation", "designation"),
    ("Description", "description"),
    ("Quantité", "qty"),
    ("Quantite", "qty"),
    ("Quantity", "qty"),
    ("Qty", "qty"),
    ("Qté", "qty"),
    ("Qt", "qty"),
    ("Qte", "qty"),
]

_SKIP_AS_HEADING_PREFIX = ("note", "attention", "warning", "caution", "important")


def _split_label_cell(cell: str) -> tuple[Optional[str], str]:
    """Split ``cell`` into (matched header label, remaining data).

    Returns (None, cell) when the cell does not *start* with a known label.
    """
    stripped = clean_text(cell)
    low = stripped.lower()
    for label, _group in _HEADER_VOCAB:
        if low == label.lower() or low.startswith(label.lower() + " "):
            return stripped[: len(label)], stripped[len(label) :].strip()
    return None, stripped


def classify_header_cell(cell: object) -> Optional[str]:
    """Return a canonical group for a header cell: item / partnumber /
    designation / description / qty, or None when it is not a header."""
    label, _rest = _split_label_cell(clean_text(cell))
    if label is None:
        return None
    low = label.lower()
    for disp, group in _HEADER_VOCAB:
        if disp.lower() == low:
            return group
    return None


# --------------------------------------------------------------------------
# Parsing
# --------------------------------------------------------------------------

_HEADING_ATX_RE = re.compile(r"^(#{1,6})\s+(.*)$")
_BOLD_HEADING_RE = re.compile(r"^\*\*.+\*\*$")
_LIST_LABEL_RE = re.compile(r"^([A-Za-z0-9]{1,3})\)\s+\S.{0,120}$")


@dataclass
class Heading:
    level: int
    text: str


@dataclass
class Table:
    """One extracted Markdown table + the component (heading) it belongs to."""

    source_file: str
    source_line: int
    component_path: list[str]  # active headings, ancestor -> leaf
    headers: list[str]         # final column labels (as printed in the doc)
    pn_col: Optional[int]      # index of the Part-Number column, if any
    rows: list[list[str]]      # cleaned data rows (already filtered / rebuilt)
    raw_row_count: int = 0     # number of visual rows seen before cleaning
    warnings: list[str] = field(default_factory=list)

    @property
    def component(self) -> str:
        return self.component_path[-1] if self.component_path else "(untitled table)"

    @property
    def is_parts_table(self) -> bool:
        return self.pn_col is not None


def _is_table_line(line: str) -> bool:
    return line.lstrip().startswith("|") and line.count("|") >= 2


def _split_row(line: str) -> list[str]:
    body = line.strip()
    if body.startswith("|"):
        body = body[1:]
    if body.endswith("|"):
        body = body[:-1]
    cells = body.split("|")
    return [clean_text(c) for c in cells]


def _is_separator_row(cells: Sequence[str]) -> bool:
    body = "".join(cells)
    return bool(body) and all(ch in " :-|" for ch in body)


def extract_tables(md_text: str, source_file: str = "") -> list[Table]:
    """Parse ``md_text`` and return every table found, tagged with its heading."""
    tables: list[Table] = []
    stack: list[Heading] = []
    pending: list[Optional[list[str]]] = []
    pending_start = 0

    def flush() -> None:
        nonlocal pending, pending_start
        if pending:
            tbl = _build_table(pending, stack, source_file, pending_start)
            if tbl is not None:
                tables.append(tbl)
            pending = []
            pending_start = 0

    lines = md_text.splitlines()
    for idx, raw in enumerate(lines, start=1):
        line = raw.rstrip()
        stripped = line.strip()
        if not stripped:
            flush()
            continue

        m = _HEADING_ATX_RE.match(line)
        if m:
            flush()
            level = len(m.group(1))
            text = clean_text(m.group(2))
            _push_heading(stack, Heading(level, text))
            continue

        if _is_table_line(line):
            cells = _split_row(line)
            if not pending:
                pending_start = idx
            if _is_separator_row(cells):
                pending.append(None)  # separator marker
            else:
                pending.append(cells)
            continue

        # --- pseudo headings (bold-only lines, "a) sub-part" labels) --------
        pseudo = _pseudo_heading(stripped)
        if pseudo is not None:
            flush()
            _push_heading(stack, pseudo)
            continue

        flush()

    flush()
    return tables


def _pseudo_heading(stripped: str) -> Optional[Heading]:
    """Convert converter-style headings that lack a '#' prefix.

    Marker keeps some component headings as bold-only lines
    (``**12) Valve …**``) or as plain "a) Sub part" labels.  They are returned
    as headings so that the following table can be named after them.
    """
    if not stripped or stripped.startswith("!"):
        return None
    if _BOLD_HEADING_RE.match(stripped):
        text = clean_text(stripped)
        if text.lower().startswith(_SKIP_AS_HEADING_PREFIX):
            return None
        return Heading(4, text)
    if _LIST_LABEL_RE.match(stripped):
        text = clean_text(stripped)
        if text.lower().startswith(_SKIP_AS_HEADING_PREFIX):
            return None
        return Heading(5, text)
    return None


def _push_heading(stack: list[Heading], heading: Heading) -> None:
    while stack and stack[-1].level >= heading.level:
        stack.pop()
    stack.append(heading)


# --------------------------------------------------------------------------
# Table assembly
# --------------------------------------------------------------------------

def _build_table(
    pending: list[Optional[list[str]]],
    stack: list[Heading],
    source_file: str,
    line_no: int,
) -> Optional[Table]:
    rows = [r for r in pending if r is not None]
    if not rows:
        return None
    if not any(any(c for c in r) for r in rows):
        return None

    path = [h.text for h in stack]

    # 1) header -------------------------------------------------------------
    header: list[str]
    data: list[list[str]]
    warnings: list[str] = []

    fused = _try_split_fused_header(rows)
    if fused is not None:
        header, first_data, rest = fused
        data = ([first_data] if first_data and any(first_data) else []) + rest
    else:
        hdr_end = 0
        while hdr_end < len(rows) and _is_header_like(rows[hdr_end]):
            hdr_end += 1
        if hdr_end == 0:
            hdr_end = 1  # no recognisable header: use the first line as header
        header = _merge_header_rows(rows[:hdr_end])
        data = rows[hdr_end:]

    if not header:
        header = [f"Column {i + 1}" for i in range(max((len(r) for r in data), default=1))]

    # 2) find the Part-Number column ----------------------------------------
    pn_col = _detect_pn_col(header, data)

    # 3) clean rows ----------------------------------------------------------
    if pn_col is not None:
        # make every row the same width as the header
        width = len(header)
        data = [list(r) + [""] * (width - len(r)) for r in data]
        data = [r[:width] for r in data]

        # relabel an ambiguous header like "Item # Part Number" -> "Part Number"
        combined = re.sub(r"[^a-z0-9]", "", header[pn_col].lower())
        if "item" in combined and "part" in combined:
            header[pn_col] = "Part Number"

        raw_row_count = len(data)
        parts = _rebuild_parts(data, pn_col, header, warnings)
        rows_out = [p["row"] for p in parts]
    else:
        # not a parts table: keep rows untouched (natural width) so that wide
        # maintenance tables are not truncated to the header width.
        maxw = max([len(header)] + [len(r) for r in data])
        if len(header) < maxw:
            header = header + [f"Column {i + 1}" for i in range(len(header), maxw)]
        data = [list(r) + [""] * (maxw - len(r)) for r in data]
        raw_row_count = len(data)
        rows_out = [r for r in data if any(c.strip() for c in r)]

    return Table(
        source_file=source_file,
        source_line=line_no,
        component_path=path,
        headers=header,
        pn_col=pn_col,
        rows=rows_out,
        raw_row_count=raw_row_count,
        warnings=warnings,
    )


def _is_header_like(row: Sequence[str]) -> bool:
    return any(classify_header_cell(c) is not None for c in row if c.strip())


def _merge_header_rows(rows: Sequence[Sequence[str]]) -> list[str]:
    width = max((len(r) for r in rows), default=0)
    merged: list[str] = []
    for col in range(width):
        label = ""
        for r in rows:
            if col < len(r) and r[col].strip():
                label = r[col]
        merged.append(clean_text(label))
    return merged


def _try_split_fused_header(rows: Sequence[Sequence[str]]) -> Optional[tuple[list[str], list[str], list[list[str]]]]:
    """Split a header row that also contains the first data row.

    e.g. ``| Item # 1 2 | Part Number 410001183677 410001183678 | … | Qty 1 1 |``
    becomes header ``[Item #, Part Number, …, Qty]`` and one data row
    ``[1 2, 410001183677 410001183678, …, 1 1]``.
    """
    first = rows[0]
    labels: list[str] = []
    leftovers: list[str] = []
    any_data = False
    for cell in first:
        label, rest = _split_label_cell(cell)
        if label is None:
            return None  # the line does not look like a labelled header
        labels.append(label)
        leftovers.append(rest)
        if rest and (has_part_number(rest) or re.search(r"(^|\s)\d{1,3}(\s|$)", rest)):
            any_data = True
    if not any_data:
        return None
    return labels, leftovers, [list(r) for r in rows[1:]]


def _detect_pn_col(header: Sequence[str], data: Sequence[Sequence[str]]) -> Optional[int]:
    # 1) exact label match, e.g. "Part Number", "PN", "Réf", "Code"
    for i, h in enumerate(header):
        if classify_header_cell(h) == "partnumber":
            return i
    # 2) a merged label like "Item # Part Number" (converter glued two headers)
    for i, h in enumerate(header):
        norm = re.sub(r"[^a-z0-9]", "", clean_text(h).lower())
        if "part" in norm and norm.startswith(("item", "no", "n", "rep", "pos")):
            return i
    return None


def _col_group(headers: Sequence[str], col: int) -> str:
    return classify_header_cell(headers[col]) or "text"


def _rebuild_parts(
    data: Sequence[Sequence[str]],
    pn_col: int,
    headers: Sequence[str],
    warnings: list[str],
) -> list[dict]:
    """Group visual rows into logical parts.

    * A new part starts on a row whose Part-Number cell is non-empty.
    * Rows without a part number are merged into the previous part when that
      part has no quantity yet (this is how wrapped text / quantities that the
      converter pushed onto a following line are recovered).
    * Rows that contain several part numbers are expanded into several parts.

    Rows that end up without a part number (e.g. "Ignition key (Customer
    provides)") are *dropped* -- the whole point of the filter is to keep only
    rows that carry a part number.
    """
    parts: list[dict] = []

    for row in data:
        tokens = pn_tokens_in(row[pn_col]) if pn_col < len(row) else []

        if len(tokens) > 1:
            # several parts collapsed onto one visual line
            for sub in _expand_fused_row(row, tokens, pn_col, headers, warnings):
                parts.append(sub)
            continue

        if len(tokens) == 1:
            parts.append({"row": list(row)})
            continue

        # ---- no part number on this row: try to merge as wrapped text -------
        if not parts:
            continue
        prev = parts[-1]["row"]
        if not _should_merge_continuation(prev, headers):
            continue  # previous part already complete; drop this orphan row
        _merge_continuation(prev, row, headers)

    return parts


def _qty_col(headers: Sequence[str]) -> Optional[int]:
    for i, h in enumerate(headers):
        if classify_header_cell(h) == "qty":
            return i
    return None


def _should_merge_continuation(part_row: Sequence[str], headers: Sequence[str]) -> bool:
    """Can the next (part-number-less) row still belong to ``part_row``?

    When a Qty column exists, a part that already carries a quantity is
    considered complete (a following no-PN row is an orphan and is dropped).
    When there is no Qty column, wrapped text is the norm, so keep merging.
    """
    q = _qty_col(headers)
    if q is not None and q < len(part_row):
        return not str(part_row[q]).strip()
    return True


def _merge_continuation(prev: list[str], row: Sequence[str], headers: Sequence[str]) -> None:
    for col, cell in enumerate(row):
        if col >= len(prev):
            break
        if not cell.strip():
            continue
        group = classify_header_cell(headers[col]) if col < len(headers) else None
        if group == "qty":
            if not prev[col].strip():
                prev[col] = cell
            continue
        prev_cell = prev[col].strip()
        if prev_cell in ("-", "–", "—", "|"):
            prev[col] = cell
        elif not prev_cell:
            prev[col] = cell
        else:
            prev[col] = prev[col] + " " + cell


def _expand_fused_row(
    row: Sequence[str],
    tokens: list[str],
    pn_col: int,
    headers: Sequence[str],
    warnings: list[str],
) -> list[dict]:
    n = len(tokens)
    out: list[dict] = []
    item_col = next((i for i, h in enumerate(headers) if classify_header_cell(h) == "item"), None)
    qty_col = next((i for i, h in enumerate(headers) if classify_header_cell(h) == "qty"), None)
    text_cols = [i for i, h in enumerate(headers) if classify_header_cell(h) in ("designation", "description", None)]

    def pieces(idx: Optional[int]) -> list[str]:
        if idx is None or idx >= len(row):
            return []
        return str(row[idx] or "").split()

    item_pieces = pieces(item_col)
    qty_pieces = pieces(qty_col)

    for i in range(n):
        new = [""] * len(row)
        new[pn_col] = tokens[i]
        if item_col is not None and len(item_pieces) == n:
            new[item_col] = item_pieces[i]
        if qty_col is not None and len(qty_pieces) == n:
            new[qty_col] = qty_pieces[i]
        for c in text_cols:
            if c >= len(row):
                continue
            fragments = _split_text_fragments(str(row[c] or ""), n)
            if len(fragments) == n:
                new[c] = fragments[i]
            else:
                if i == 0:
                    new[c] = str(row[c] or "")
                warnings.append(
                    f"'{headers[c] if c < len(headers) else c}': "
                    f"could not split fused text into {n} parts (got {len(fragments)})"
                )
        out.append({"row": new})
    if n > 1:
        warnings.insert(0, f"line contained {n} part numbers that were expanded on a best-effort basis")
    return out


def _split_text_fragments(text: str, n: int) -> list[str]:
    """Split concatenated column text into ``n`` fragments, best effort.

    Works well when every fragment starts with a capital letter (which is the
    pattern in the PDF conversion output) and degrades to an even split.
    """
    tokens = text.split()
    if not tokens or n <= 1:
        return [text]
    if len(tokens) < n:
        return [text]  # not enough words; caller will flag it

    # candidate boundaries: before an uppercase word or an opening bracket
    candidates = [
        i
        for i in range(1, len(tokens))
        if tokens[i][0].isupper() or tokens[i][0] in "(«\"'#0123456789"
    ]

    # choose n-1 boundaries that are as evenly spread as possible
    ideal = [(k * len(tokens)) / n for k in range(1, n)]
    chosen: list[int] = []
    for target in ideal:
        usable = [c for c in candidates if c > (chosen[-1] if chosen else 0)]
        if not usable:
            break
        best = min(usable, key=lambda c: abs(c - target))
        chosen.append(best)
        if len(chosen) == n - 1:
            break

    if len(chosen) < n - 1:
        # fall back to an even split
        chosen = sorted({round(k * len(tokens) / n) for k in range(1, n)})
        chosen = [max(1, min(len(tokens) - 1, c)) for c in chosen]

    fragments: list[str] = []
    start = 0
    for end in sorted(chosen) + [len(tokens)]:
        fragments.append(" ".join(tokens[start:end]))
        start = end
    return fragments


# --------------------------------------------------------------------------
# Excel export
# --------------------------------------------------------------------------

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _sheet_name(component: str, used: set[str]) -> str:
    name = _INVALID_SHEET_CHARS.sub("-", component)
    name = _WS_RE.sub(" ", name).strip(" -")
    if not name:
        name = "Table"
    name = name[:31].rstrip(" -")
    base, i = name, 2
    while name.lower() in used:
        suffix = f" {i}"
        name = f"{base[: 31 - len(suffix)]}{suffix}"
        i += 1
    used.add(name.lower())
    return name


_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_TITLE_FILL = PatternFill("solid", fgColor="DDEBF7")
_NOTE_FONT = Font(color="9C5700", italic=True, size=9)
_THIN = Side(style="thin", color="B7C3D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _autosize(ws) -> None:
    from openpyxl.cell.cell import MergedCell

    for col_cells in ws.columns:
        vals = [c.value for c in col_cells if not isinstance(c, MergedCell)]
        width = max((len(str(v or "")) for v in vals), default=8)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 60)


def export_tables_to_excel(
    tables: Iterable[Table],
    out_path: str | Path,
    *,
    include_other_tables: bool = False,
    write_index_sheet: bool = True,
) -> dict:
    """Write the given tables into an Excel workbook.

    * one worksheet per component table (sheet title = component heading)
    * only parts tables (a Part Number column exists) by default; pass
      ``include_other_tables=True`` to also export the tables that have no
      dedicated Part Number column (e.g. maintenance schedules).
    * every row that does not carry a Part Number is dropped.
    """
    tables = list(tables)
    parts = [t for t in tables if t.is_parts_table]
    others = [t for t in tables if not t.is_parts_table]

    selected = parts if not include_other_tables else tables

    wb = Workbook()
    wb.remove(wb.active)

    used: set[str] = set()
    index_rows: list[dict] = []

    for t in selected:
        ws = wb.create_sheet(_sheet_name(t.component, used))

        # --- component header rows (the table's heading) -------------------
        ncols = max(len(t.headers), 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        title = ws.cell(row=1, column=1, value=t.component)
        title.font = Font(bold=True, size=13, color="1F4E79")
        title.alignment = Alignment(horizontal="left", vertical="center")
        title.fill = _TITLE_FILL

        if len(t.component_path) > 1:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
            sec = ws.cell(row=2, column=1, value="Section: " + "  >  ".join(t.component_path[:-1]))
            sec.font = Font(size=9, color="44546A", italic=True)

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
        src = ws.cell(row=3, column=1, value=f"Source: {t.source_file}  (line {t.source_line})")
        src.font = Font(size=8, color="808080")

        # --- table column headers ------------------------------------------
        hr = 4
        for c, label in enumerate(t.headers, start=1):
            cell = ws.cell(row=hr, column=c, value=label)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = _HEADER_FILL
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.freeze_panes = ws.cell(row=hr + 1, column=1)

        # --- data rows ------------------------------------------------------
        for r_i, row in enumerate(t.rows, start=hr + 1):
            for c_i in range(ncols):
                val = row[c_i] if c_i < len(row) else ""
                cell = ws.cell(row=r_i, column=c_i + 1, value=val or None)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = _BORDER

        note_row = hr + len(t.rows) + 2
        if t.pn_col is None:
            ws.cell(row=note_row, column=1, value="Note: table has no dedicated Part Number column.").font = _NOTE_FONT
        if t.warnings:
            ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ncols)
            msg = "Note: " + "; ".join(dict.fromkeys(t.warnings))
            ws.cell(row=note_row, column=1, value=msg).font = _NOTE_FONT

        _autosize(ws)
        index_rows.append(
            {
                "Component": t.component,
                "Section": "  >  ".join(t.component_path[:-1]) if t.component_path else "",
                "Worksheet": ws.title,
                "Part rows": len(t.rows),
                "Raw rows": t.raw_row_count,
                "Source": f"{t.source_file}:{t.source_line}",
            }
        )

    # --- optional index / read-me sheet -------------------------------------
    if write_index_sheet:
        ws = wb.create_sheet("Index", 0)
        headers = ["Component", "Section", "Worksheet", "Part rows", "Raw rows", "Source"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = _HEADER_FILL
        for r_i, row in enumerate(index_rows, start=2):
            for c_i, key in enumerate(headers, start=1):
                cell = ws.cell(row=r_i, column=c_i, value=row.get(key) or "")
                cell.border = _BORDER
        ws.cell(row=len(index_rows) + 2, column=1,
                value="Legend: each worksheet = one component (its table heading is shown on the first row). "
                      "Only rows carrying a Part Number are exported.").font = _NOTE_FONT
        ws.freeze_panes = "A2"
        _autosize(ws)

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    return {
        "path": str(out),
        "tables": len(selected),
        "parts_tables": len(parts),
        "other_tables": len(others),
        "parts_rows": sum(len(t.rows) for t in parts),
        "workbook_sheets": len(wb.sheetnames),
    }
