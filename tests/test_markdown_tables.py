"""Unit tests for src/ingressor/markdown_tables.py

Run with:  python -m unittest discover -s tests -v
"""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from ingressor.markdown_tables import (  # noqa: E402
    clean_text,
    extract_tables,
    export_tables_to_excel,
    has_part_number,
    is_part_number_token,
)


SAMPLE = """\
# Machine manual

## 1) Cooler assembly

Some intro text.

| Item # | Part Number | Désignation           | Description            | Qty |
|--------|-------------|------------------------|------------------------|-----|
| 1      | A08-0007W   | Refroidisseur hydraul. | Hydraulic cooler       | 1   |
| 2      | A04-0007T   | Valve de freinage      | Brake control valve    |     |
|        |             |                        | (with brake pedal)     | 1   |
|        | A09014Q     | Kit de joints          | Seal kit               |     |
| 3      | 510000108714| Affichage couleur      | color display          | 1   |

Note: the valve part above has no part number and must be dropped.

## 2) Driver station panel

| Item # 1 2 | Part Number 410001183677 410001183678 | Désignation Console panel A Console panel B | Description Panel A Panel B | Qty 1 1 |
|------------|---------------------------------------|---------------------------------------------|-----------------------------|---------|
| 3          | 510000108714                          | Display                                    | color display               | 1       |
"""


class TokenTests(unittest.TestCase):
    def test_real_part_numbers(self):
        for pn in (
            "410001183677",
            "058773M",
            "A02-0003K",
            "A01092Y",
            "P04-0008H",
            "T15004-SGP00702",
            "487327J",
            "003088V",
            "X22124F",
            "A09082N/00",
        ):
            self.assertTrue(is_part_number_token(pn), pn)

    def test_not_part_numbers(self):
        for token in ("1", "14", "397", "10000", "3.5m", "X22124F/00.......", "Oil", "250 h"):
            self.assertFalse(is_part_number_token(token), token)


class CleanTests(unittest.TestCase):
    def test_strips_tags_and_emphasis(self):
        self.assertEqual(clean_text("<b>Engine</b> **swap**  part"), "Engine swap part")


class ExtractTests(unittest.TestCase):
    def setUp(self):
        self.tables = extract_tables(SAMPLE, source_file="sample.md")

    def test_tables_found(self):
        self.assertEqual(len(self.tables), 2)

    def test_component_heading(self):
        self.assertEqual(self.tables[0].component, "1) Cooler assembly")
        self.assertEqual(self.tables[1].component, "2) Driver station panel")

    def test_pn_column_detected(self):
        self.assertIsNotNone(self.tables[0].pn_col)
        self.assertEqual(self.tables[0].headers[self.tables[0].pn_col], "Part Number")

    def test_wrapped_rows_merged_and_no_pn_rows_dropped(self):
        t = self.tables[0]
        rows = {row[t.pn_col]: row for row in t.rows}
        # wrapped description + qty recovered on the A04-0007T part
        valve = rows.get("A04-0007T")
        self.assertIsNotNone(valve)
        desc = " ".join(valve).lower()
        self.assertIn("with brake pedal", desc)
        # the "Seal kit" part and the display part are separate rows
        self.assertIn("A09014Q", rows)
        self.assertIn("510000108714", rows)
        # a row with no part number must not be exported
        self.assertEqual(len(t.rows), 4)  # 1,2(+merge), seal kit, display... actually count below
        self.assertFalse(any("no part number" in c.lower() for r in t.rows for c in r))

    def test_fused_header_expanded(self):
        t = self.tables[1]
        self.assertEqual(t.pn_col, 1)
        pns = [row[1] for row in t.rows]
        self.assertEqual(pns[:2], ["410001183677", "410001183678"])
        self.assertEqual(len(t.rows), 3)  # items 1,2 + item 3

    def test_plain_table_without_pn_column_is_not_parts_table(self):
        md = "# T\n\n| System | 250 h | 500 h |\n|---|---|---|\n| Engine | oil | filter |\n"
        t = extract_tables(md, source_file="x")[0]
        self.assertIsNone(t.pn_col)
        self.assertFalse(t.is_parts_table)


class ExportTests(unittest.TestCase):
    def test_export_writes_workbook(self):
        from openpyxl import load_workbook

        tables = extract_tables(SAMPLE, source_file="sample.md")
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "parts.xlsx"
            summary = export_tables_to_excel(tables, out)
            self.assertTrue(out.exists())
            wb = load_workbook(out)
            # Index + 2 part tables
            self.assertEqual(len(wb.sheetnames), 3)
            # first row of each part sheet = component heading
            ws = wb[summary["path"] and "2) Driver station panel"]
            self.assertEqual(ws.cell(row=1, column=1).value, "2) Driver station panel")
            self.assertIsNotNone(ws.freeze_panes)


class RealFixtureTests(unittest.TestCase):
    """End-to-end smoke test against the real servicing manual if present."""

    FIXTURE = Path(__file__).resolve().parent.parent / (
        "conversion_results/1. Servicing Manual TAM 7002.pdf 01072026/"
        "1. Servicing Manual TAM 7002.pdf 01072026.md"
    )

    def test_real_manual_extracts_parts(self):
        if not self.FIXTURE.exists():
            self.skipTest("fixture markdown not available")
        tables = extract_tables(self.FIXTURE.read_text(encoding="utf-8"), str(self.FIXTURE))
        parts = [t for t in tables if t.is_parts_table]
        self.assertGreater(len(parts), 40)
        self.assertGreater(sum(len(t.rows) for t in parts), 300)


if __name__ == "__main__":
    unittest.main()
